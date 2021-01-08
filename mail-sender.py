# Program 3 z 10 w ramach Ferie Challange 2021

# Stwórz prosty program, który będzie wysyłał spersonalizowany mailing do wybranych osób. “Bazą danych” 
# jest plik Excela (aby było “ciekawiej” 😉 ) lub CSV, zawierający dwie kolumny z nagłówkami: “E-mail” oraz 
# “Imię i nazwisko” (zakładamy, że zawsze w takiej kolejności, a imię i nazwisko są oddzielone spacją). 
# Do użytkowników należy wysłać maila z tematem “Your image” oraz spersonalizowaną prostą treścią np. “Hi {Imię}! 
# it’s file generated for you”. Dodatkowo w załączniku maila znajduje się plik graficzny o nazwie 
# “{Imię}_{Nazwisko}_image.png” (pliki są w zadanej lokalizacji). 
# Odpowiednio zabezpiecz program (np. brakujący plik Excela, brakujące dane w Excelu, brak pliku png) 
# oraz zabezpiecz przed spamowaniem (np. jeden mail wysyłany co 1 sekundę). Mogą przydać się moduły: 
# smtplib, email, ssl, xlrd, re, os. 

# Propozycje rozszerzenia: dodaj opcję wysyłania maili z treścią w HTML oraz walidator poprawności maila 
# (np. używając wyrażeń regularnych - moduł re).

import email 
import smtplib
import ssl
import getpass
import imghdr
import os
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl import load_workbook

# Getting data from an excel file

try:
    wb = load_workbook('mail.xlsx')
    ws = wb.get_sheet_by_name("Arkusz1")
except Exception:
    print("The file doesn't exsit")

data = []

for row in ws.iter_rows(values_only=True):
    data.append(row)

data.remove(data[0])

# logging in to the e-mail

port = 465  # For SSL
smtp_server = "smtp.gmail.com"
sender_email = "programistka.python@gmail.com"
password = getpass.getpass(prompt='Type your password and press enter: ', stream=None)

for i in data:
    try:
        d = i[1].split(" ")
        message = EmailMessage()
        message["Subject"] = f"Your image"
        message["From"] = sender_email
        message["To"] = i[0]
        message.set_content(f"Hi {d[0]}! It's file generated for you")
    except Exception:
        print(f'Some data missin in row {data.index(i)+2}')

text = message.get_content().as_string()

# Adding an attachment

image = f"{d[0]}_{d[1]}_image.png"

try:
    with open(image, "rb") as pi:
        img_data = pi.read()
except Exception:
    print("The image doesn't exist or the wrong name was entered")
message.add_attachment(img_data, maintype="image", subtype=imghdr.what(None, img_data)) 


# Create a secure SSL context & send mail
context = ssl.create_default_context()

with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
    server.login(sender_email, password)
    server.sendmail(sender_email, message["From"], text)

